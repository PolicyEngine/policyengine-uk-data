"""OBR Economic and Fiscal Outlook targets.

Downloads and parses the OBR's detailed supplementary tables (receipts
and expenditure xlsx) to extract tax receipt forecasts, benefit
expenditure, and benefit caseloads.

Sources:
- Receipts: https://obr.uk/download/november-2025-economic-and-fiscal-outlook-detailed-forecast-tables-receipts/
- Expenditure: https://obr.uk/download/november-2025-economic-and-fiscal-outlook-detailed-forecast-tables-expenditure/
"""

import io
import logging
from functools import lru_cache
from pathlib import Path

import openpyxl
import requests
import yaml

from policyengine_uk_data.targets.schema import Target, Unit

logger = logging.getLogger(__name__)

_SOURCES_YAML = Path(__file__).parent.parent / "sources.yaml"

# Financial year columns in OBR tables: C=2024-25, D=2025-26, ..., I=2030-31
# PolicyEngine convention: FY 2025-26 → calendar year 2025 (first year)
_FY_COL_TO_YEAR = {
    "C": 2024,
    "D": 2025,
    "E": 2026,
    "F": 2027,
    "G": 2028,
    "H": 2029,
    "I": 2030,
}

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
    ),
}


def _load_config():
    with open(_SOURCES_YAML) as f:
        return yaml.safe_load(f)


@lru_cache(maxsize=1)
def _download_workbook(url: str) -> openpyxl.Workbook:
    """Download an xlsx from OBR and return an openpyxl workbook."""
    r = requests.get(url, headers=_HEADERS, allow_redirects=True, timeout=60)
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)


def _read_row_values(
    ws, row_num: int, col_letters: list[str]
) -> dict[int, float]:
    """Read numeric values from a row, mapped to calendar years."""
    result = {}
    for col in col_letters:
        cell = ws[f"{col}{row_num}"]
        val = cell.value
        if val is not None and isinstance(val, (int, float)):
            result[_FY_COL_TO_YEAR[col]] = float(val) * 1e9
    return result


def _find_row(ws, label: str, col: str = "B", max_row: int = 80) -> int:
    """Find the row number where a cell starts with label."""
    for row in range(1, max_row + 1):
        cell_val = ws[f"{col}{row}"].value
        if cell_val and str(cell_val).strip().startswith(label):
            return row
    raise ValueError(f"Row '{label}' not found in sheet")


def _parse_receipts(wb: openpyxl.Workbook) -> list[Target]:
    """Parse tax receipts from the OBR EFO.

    Income tax uses Table 3.4 (accrued basis) for consistency with
    the standard fiscal forecasting convention. Other receipts use
    Table 3.9 (cash basis) since they only appear there.
    """
    config = _load_config()
    vintage = config["obr"]["vintage"]
    ref = config["obr"]["efo_receipts"]
    cols_34 = list(_FY_COL_TO_YEAR.keys())

    # Table 3.9 columns are shifted right by one vs 3.4
    cols_39 = ["D", "E", "F", "G", "H", "I", "J"]
    fy_39 = {
        "D": 2024,
        "E": 2025,
        "F": 2026,
        "G": 2027,
        "H": 2028,
        "I": 2029,
        "J": 2030,
    }

    def read_39(ws, row_num: int) -> dict[int, float]:
        result = {}
        for col in cols_39:
            cell = ws[f"{col}{row_num}"]
            val = cell.value
            if val is not None and isinstance(val, (int, float)):
                result[fy_39[col]] = float(val) * 1e9
        return result

    targets = []

    # Income tax from Table 3.4 (accrued basis)
    try:
        ws34 = wb["3.4"]
        row_num = _find_row(
            ws34, "Income tax (gross of tax credits)", col="B", max_row=30
        )
        values = _read_row_values(ws34, row_num, cols_34)
        if values:
            targets.append(
                Target(
                    name="obr/income_tax",
                    variable="income_tax",
                    source="obr",
                    unit=Unit.GBP,
                    values=values,
                    reference_url=ref,
                    forecast_vintage=vintage,
                )
            )
    except ValueError:
        logger.warning("OBR receipts: income tax row not found in 3.4")

    # Other receipts from Table 3.9 (cash basis)
    ws39 = wb["3.9"]
    cash_rows = {
        "ni": ("National insurance contributions", "ni_employee"),
        "vat": ("Value added tax", "vat"),
        "fuel_duties": ("Fuel duties", "fuel_duty"),
        "capital_gains_tax": ("Capital gains tax", "capital_gains_tax"),
        "sdlt": ("Stamp duty land tax", "stamp_duty_land_tax"),
    }

    for name, (label, variable) in cash_rows.items():
        try:
            row_num = _find_row(ws39, label, col="B", max_row=80)
            values = read_39(ws39, row_num)
            if values:
                targets.append(
                    Target(
                        name=f"obr/{name}",
                        variable=variable,
                        source="obr",
                        unit=Unit.GBP,
                        values=values,
                        reference_url=ref,
                        forecast_vintage=vintage,
                    )
                )
        except ValueError:
            logger.warning("OBR receipts: row '%s' not found", label)

    return targets


def _parse_council_tax(wb: openpyxl.Workbook) -> list[Target]:
    """Parse Table 4.1 (council tax receipts) from expenditure xlsx."""
    config = _load_config()
    vintage = config["obr"]["vintage"]
    ref = config["obr"]["efo_expenditure"]
    ws = wb["4.1"]

    cols = ["C", "D", "E", "F", "G", "H", "I"]
    fy = {
        "C": 2024,
        "D": 2025,
        "E": 2026,
        "F": 2027,
        "G": 2028,
        "H": 2029,
        "I": 2030,
    }

    def read_41(row_num: int) -> dict[int, float]:
        result = {}
        for col in cols:
            cell = ws[f"{col}{row_num}"]
            val = cell.value
            if val is not None and isinstance(val, (int, float)):
                result[fy[col]] = float(val) * 1e9
        return result

    ct_rows = {
        "council_tax": ("Total net council tax receipts", "council_tax"),
        "council_tax_england": (
            "England council tax receipts",
            "council_tax",
        ),
        "council_tax_scotland": (
            "Scotland council tax receipts",
            "council_tax",
        ),
        "council_tax_wales": ("Wales council tax receipts", "council_tax"),
        "domestic_rates": ("NI domestic rates", "domestic_rates"),
    }

    targets = []
    for name, (label, variable) in ct_rows.items():
        try:
            row_num = _find_row(ws, label, col="B", max_row=30)
            values = read_41(row_num)
            if values:
                targets.append(
                    Target(
                        name=f"obr/{name}",
                        variable=variable,
                        source="obr",
                        unit=Unit.GBP,
                        values=values,
                        reference_url=ref,
                        forecast_vintage=vintage,
                    )
                )
        except ValueError:
            logger.warning("OBR council tax: row '%s' not found", label)

    return targets


def _parse_nics(wb: openpyxl.Workbook) -> list[Target]:
    """Parse Table 3.4 (income tax and NICs detail) for employee/employer."""
    config = _load_config()
    vintage = config["obr"]["vintage"]
    ref = config["obr"]["efo_receipts"]
    ws = wb["3.4"]
    cols = list(_FY_COL_TO_YEAR.keys())

    nic_rows = {
        "ni_employee": (
            "Class 1 Employee NICs",
            "ni_employee",
        ),
        "ni_employer": (
            "Class 1 Employer NICs",
            "ni_employer",
        ),
    }

    targets = []
    for name, (label, variable) in nic_rows.items():
        try:
            row_num = _find_row(ws, label, col="B", max_row=30)
            values = _read_row_values(ws, row_num, cols)
            if values:
                targets.append(
                    Target(
                        name=f"obr/{name}",
                        variable=variable,
                        source="obr",
                        unit=Unit.GBP,
                        values=values,
                        reference_url=ref,
                        forecast_vintage=vintage,
                    )
                )
        except ValueError:
            logger.warning("OBR NICs: row '%s' not found", label)

    return targets


def _parse_welfare(wb: openpyxl.Workbook) -> list[Target]:
    """Parse Table 4.9 (welfare spending) from expenditure xlsx."""
    config = _load_config()
    vintage = config["obr"]["vintage"]
    ref = config["obr"]["efo_expenditure"]
    ws = wb["4.9"]

    cols = ["C", "D", "E", "F", "G", "H", "I"]
    fy = {
        "C": 2024,
        "D": 2025,
        "E": 2026,
        "F": 2027,
        "G": 2028,
        "H": 2029,
        "I": 2030,
    }

    def read_49(row_num: int) -> dict[int, float]:
        result = {}
        for col in cols:
            cell = ws[f"{col}{row_num}"]
            val = cell.value
            if val is not None and isinstance(val, (int, float)):
                result[fy[col]] = float(val) * 1e9
        return result

    benefit_rows = {
        "housing_benefit": (
            "Housing benefit (not on JSA)",
            "housing_benefit",
        ),
        "pip": (
            "Disability living allowance and personal independence p",
            "pip",
        ),
        "esa": ("Incapacity benefits", "esa_income"),
        "attendance_allowance": (
            "Attendance allowance",
            "attendance_allowance",
        ),
        "pension_credit": ("Pension credit", "pension_credit"),
        "carers_allowance": ("Carer's allowance", "carers_allowance"),
        "statutory_maternity_pay": (
            "Statutory maternity pay",
            "statutory_maternity_pay",
        ),
        "winter_fuel_allowance": (
            "Winter fuel payment",
            "winter_fuel_allowance",
        ),
        "universal_credit_in_cap": (
            "Universal credit",
            "universal_credit",
        ),
        "child_benefit": ("Child benefit", "child_benefit"),
        "state_pension": ("State pension", "state_pension"),
        "jobseekers_allowance": (
            "Jobseeker's allowance",
            "jsa_income",
        ),
    }

    targets = []
    # Welfare cap section (rows 6-36)
    for name, (label, variable) in benefit_rows.items():
        try:
            row_num = _find_row(ws, label, col="B", max_row=55)
            values = read_49(row_num)
            if values:
                targets.append(
                    Target(
                        name=f"obr/{name}",
                        variable=variable,
                        source="obr",
                        unit=Unit.GBP,
                        values=values,
                        reference_url=ref,
                        forecast_vintage=vintage,
                    )
                )
        except ValueError:
            logger.warning("OBR welfare: row '%s' not found", label)

    # Universal credit outside cap (row 43) is jobseekers UC
    try:
        # UC outside cap = predominantly JSA-conditionality UC
        uc_outside_row = _find_row(ws, "Universal credit", col="B", max_row=55)
        # Find the second UC row (outside cap section)
        for row in range(uc_outside_row + 1, 55):
            cell_val = ws[f"B{row}"].value
            if cell_val and str(cell_val).strip().startswith(
                "Universal credit"
            ):
                values = read_49(row)
                if values:
                    targets.append(
                        Target(
                            name="obr/universal_credit_outside_cap",
                            variable="universal_credit",
                            source="obr",
                            unit=Unit.GBP,
                            values=values,
                            reference_url=ref,
                            forecast_vintage=vintage,
                        )
                    )
                break
    except ValueError:
        logger.warning("OBR welfare: UC outside cap not found")

    return targets


def _parse_tv_licence(wb: openpyxl.Workbook) -> list[Target]:
    """Parse Table 4.19 (BBC) from expenditure xlsx."""
    config = _load_config()
    vintage = config["obr"]["vintage"]
    ref = config["obr"]["efo_expenditure"]

    try:
        ws = wb["4.19"]
        cols = ["C", "D", "E", "F", "G", "H", "I"]
        fy = {
            "C": 2024,
            "D": 2025,
            "E": 2026,
            "F": 2027,
            "G": 2028,
            "H": 2029,
            "I": 2030,
        }

        # Find "Licence fee receipts" or "BBC licence fee"
        for row_num in range(1, 30):
            val = ws[f"B{row_num}"].value
            if val and "licence fee" in str(val).lower():
                values = {}
                for col in cols:
                    cell = ws[f"{col}{row_num}"]
                    v = cell.value
                    if v is not None and isinstance(v, (int, float)):
                        values[fy[col]] = float(v) * 1e9
                if values:
                    return [
                        Target(
                            name="obr/tv_licence_fee",
                            variable="tv_licence",
                            source="obr",
                            unit=Unit.GBP,
                            values=values,
                            reference_url=ref,
                            forecast_vintage=vintage,
                        )
                    ]
    except Exception:
        logger.warning("OBR: TV licence table not found")

    return []


# ISC census: private school students (roughly constant at ~557k)
_PRIVATE_SCHOOL = {y: 557_000 for y in range(2018, 2032)}

# SPP Review: salary sacrifice NI relief (uprated 3% pa from 2024 base)
_SS_EMPLOYEE_NI = {
    y: 1.2e9 * 1.03 ** max(0, y - 2024) for y in range(2024, 2032)
}
_SS_EMPLOYER_NI = {
    y: 2.9e9 * 1.03 ** max(0, y - 2024) for y in range(2024, 2032)
}


def get_targets() -> list[Target]:
    config = _load_config()
    targets = []

    try:
        receipts_wb = _download_workbook(config["obr"]["efo_receipts"])
        targets.extend(_parse_receipts(receipts_wb))
        targets.extend(_parse_nics(receipts_wb))
    except Exception as e:
        logger.error("Failed to download/parse OBR receipts: %s", e)

    try:
        expenditure_wb = _download_workbook(config["obr"]["efo_expenditure"])
        targets.extend(_parse_council_tax(expenditure_wb))
        targets.extend(_parse_welfare(expenditure_wb))
        targets.extend(_parse_tv_licence(expenditure_wb))
    except Exception as e:
        logger.error("Failed to download/parse OBR expenditure: %s", e)

    # Static targets that don't come from the xlsx
    targets.append(
        Target(
            name="obr/private_school_students",
            variable="attends_private_school",
            source="obr",
            unit=Unit.COUNT,
            values=_PRIVATE_SCHOOL,
            is_count=True,
            reference_url="https://www.isc.co.uk/research/annual-census/",
        )
    )
    targets.append(
        Target(
            name="obr/salary_sacrifice_employee_ni_relief",
            variable="ni_employee",
            source="obr",
            unit=Unit.GBP,
            values=_SS_EMPLOYEE_NI,
            reference_url="https://assets.publishing.service.gov.uk/media/67ce0e7c08e764d17a5d3c21/2025_SPP_Review.pdf",
        )
    )
    targets.append(
        Target(
            name="obr/salary_sacrifice_employer_ni_relief",
            variable="ni_employer",
            source="obr",
            unit=Unit.GBP,
            values=_SS_EMPLOYER_NI,
            reference_url="https://assets.publishing.service.gov.uk/media/67ce0e7c08e764d17a5d3c21/2025_SPP_Review.pdf",
        )
    )

    return targets
