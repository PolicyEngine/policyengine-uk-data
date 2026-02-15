"""ONS families & households targets.

Downloads Table 7 from the ONS Families and Households dataset to
get household counts by type (one-person, couples, lone parents, etc).

Source: https://www.ons.gov.uk/peoplepopulationandcommunity/birthsdeathsandmarriages/families/datasets/familiesandhouseholdsfamiliesandhouseholds
"""

import io
import logging
from functools import lru_cache

import openpyxl
import requests

from policyengine_uk_data.targets.schema import Target, Unit

logger = logging.getLogger(__name__)

_URL = (
    "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/"
    "birthsdeathsandmarriages/families/datasets/"
    "familiesandhouseholdsfamiliesandhouseholds/"
    "current/familiesandhouseholdsuk2024.xlsx"
)
_REF = (
    "https://www.ons.gov.uk/peoplepopulationandcommunity/"
    "birthsdeathsandmarriages/families/datasets/"
    "familiesandhouseholdsfamiliesandhouseholds"
)
_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36"
    ),
}

# Table 7 rows: (row_number, target_name)
# Row numbers are 1-indexed in the xlsx
_TABLE7_ROWS = {
    14: "lone_households_under_65",
    15: "lone_households_over_65",
    16: "unrelated_adult_households",
    19: "couple_no_children_households",
    20: "couple_under_3_children_households",
    21: "couple_3_plus_children_households",
    22: "couple_non_dependent_children_only_households",
    24: "lone_parent_dependent_children_households",
    25: "lone_parent_non_dependent_children_households",
    26: "multi_family_households",
}

# Years we want (columns follow pattern: year_col, cv_col, ci_col,
# repeating every 3 columns from col 2 for year 1996)
_MIN_YEAR = 2018


@lru_cache(maxsize=1)
def _download_workbook() -> openpyxl.Workbook:
    r = requests.get(
        _URL, headers=_HEADERS, allow_redirects=True, timeout=60
    )
    r.raise_for_status()
    return openpyxl.load_workbook(
        io.BytesIO(r.content), data_only=True
    )


def _find_year_columns(ws) -> dict[int, int]:
    """Map calendar year -> column index for Estimate columns in Table 7."""
    year_cols = {}
    for col in range(2, ws.max_column + 1):
        header = ws.cell(row=12, column=col).value
        if header and "Estimate" in str(header):
            year_str = str(header).split(" ")[0]
            try:
                year = int(year_str)
                if year >= _MIN_YEAR:
                    year_cols[year] = col
            except ValueError:
                continue
    return year_cols


def get_targets() -> list[Target]:
    targets = []
    try:
        wb = _download_workbook()
        ws = wb["7"]
        year_cols = _find_year_columns(ws)

        for row_num, name in _TABLE7_ROWS.items():
            values = {}
            for year, col in year_cols.items():
                val = ws.cell(row=row_num, column=col).value
                if val is not None and isinstance(val, (int, float)):
                    values[year] = float(val) * 1e3  # thousands → count
            if values:
                targets.append(
                    Target(
                        name=f"ons/{name}",
                        variable="family_type",
                        source="ons",
                        unit=Unit.COUNT,
                        values=values,
                        is_count=True,
                        reference_url=_REF,
                    )
                )

    except Exception as e:
        logger.error("Failed to download ONS households xlsx: %s", e)

    return targets
