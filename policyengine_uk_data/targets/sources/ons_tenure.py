"""ONS subnational dwelling stock by tenure targets.

Downloads the ONS SPREE tenure estimates to get England-level tenure
breakdowns (owned outright, owned with mortgage, private rent, social
rent) summed across all local authorities.

Source: https://www.ons.gov.uk/peoplepopulationandcommunity/housing/datasets/subnationaldwellingstockbytenureestimates
"""

import io
import logging
from functools import lru_cache

import openpyxl
import requests

from policyengine_uk_data.targets.schema import (
    GeographicLevel,
    Target,
    Unit,
)
from policyengine_uk_data.targets.sources._common import HEADERS

logger = logging.getLogger(__name__)

_URL = (
    "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/"
    "housing/datasets/subnationaldwellingstockbytenureestimates/"
    "current/subnationaldwellingsbytenure2024.xlsx"
)
_REF = (
    "https://www.ons.gov.uk/peoplepopulationandcommunity/"
    "housing/datasets/subnationaldwellingstockbytenureestimates"
)

# Tenure categories in the xlsx header → target name suffix
_TENURE_COLS = {
    "Owned Outright": "tenure_england_owned_outright",
    "Owned with Mortgage or Loan": "tenure_england_owned_with_mortgage",
    "Private Rent": "tenure_england_rented_privately",
    "Social Rent": "tenure_england_social_rent",
    "Total Dwellings": "tenure_england_total",
}


@lru_cache(maxsize=1)
def _download_workbook() -> openpyxl.Workbook:
    r = requests.get(_URL, headers=HEADERS, allow_redirects=True, timeout=60)
    r.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)


def _parse_header_columns(ws) -> dict[tuple[int, str], int]:
    """Map (year, tenure_category) → column index from row 4 headers."""
    mapping = {}
    for col in range(5, ws.max_column + 1):
        header = ws.cell(row=4, column=col).value
        if not header:
            continue
        header = str(header)
        for tenure_suffix in _TENURE_COLS:
            if header.endswith(tenure_suffix):
                year = int(header.split(" ")[0])
                mapping[(year, tenure_suffix)] = col
                break
    return mapping


def get_targets() -> list[Target]:
    targets = []
    try:
        wb = _download_workbook()
        ws = wb["1a"]
        col_map = _parse_header_columns(ws)

        # Sum across all local authorities for each (year, tenure)
        totals: dict[tuple[int, str], float] = {}
        for row in range(5, ws.max_row + 1):
            for (year, tenure), col in col_map.items():
                val = ws.cell(row=row, column=col).value
                if val is not None and isinstance(val, (int, float)):
                    key = (year, tenure)
                    totals[key] = totals.get(key, 0) + float(val)

        # Build targets
        for tenure_col, target_name in _TENURE_COLS.items():
            values = {}
            for (year, tenure), total in totals.items():
                if tenure == tenure_col:
                    values[year] = total
            if values:
                targets.append(
                    Target(
                        name=f"ons/{target_name}",
                        variable="tenure_type",
                        source="ons",
                        unit=Unit.COUNT,
                        geographic_level=GeographicLevel.COUNTRY,
                        geo_code="E",
                        geo_name="England",
                        values=values,
                        is_count=True,
                        reference_url=_REF,
                    )
                )

    except Exception as e:
        logger.error("Failed to download ONS tenure data: %s", e)

    return targets
