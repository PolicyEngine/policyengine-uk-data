"""VOA council tax band targets.

Council tax band counts (A-H + total) by region from the latest VOA
stock-of-properties summary workbook, plus Scotland from the latest
Scottish Government chargeable-dwellings workbook.

Sources:
- https://www.gov.uk/government/statistics/council-tax-stock-of-properties-2025
- https://www.gov.scot/publications/council-tax-datasets/
"""

from functools import lru_cache
from html import unescape
from io import BytesIO
import logging
import re
import time
from urllib.parse import urljoin
from zipfile import BadZipFile

import openpyxl
import requests

from policyengine_uk_data.targets.schema import (
    GeographicLevel,
    Target,
    Unit,
)
from policyengine_uk_data.targets.sources._common import HEADERS, load_config

_SHEET = "CTSOP2.0"
_HEADER_ROW = 5
_BANDS = ["A", "B", "C", "D", "E", "F", "G", "H"]
logger = logging.getLogger(__name__)
_SCOTLAND_REF = "https://www.gov.scot/publications/council-tax-datasets/"
_SCOTLAND_WORKBOOK_URL = (
    "https://www.gov.scot/binaries/content/documents/govscot/publications/"
    "statistics/2019/04/council-tax-datasets/documents/"
    "number-of-chargeable-dwellings/chargeable-dwellings---september-2025-data/"
    "chargeable-dwellings---september-2025-data/govscot%3Adocument/"
    "CTAXBASE%2B2025%2B-%2BTables%2B-%2BChargeable%2BDwellings.xlsx"
)
_SCOTLAND_WORKBOOK_LINK = re.compile(
    r'href="([^"]*chargeable-dwellings---september-2025-data[^"]+?\.xlsx)"',
    re.IGNORECASE,
)
_SCOTLAND_FALLBACK_TOTAL_2025 = 2_623_149
_SCOTLAND_FALLBACK_BAND_SHARES_2024_25 = {
    # Shares published by the Scottish Government in council-tax reform
    # analysis from the CTAXBASE council tax datasets. They are used only
    # when gov.scot's CloudFront challenge blocks the workbook download.
    "A": 0.191,
    "B": 0.223,
    "C": 0.163,
    "D": 0.140,
    "E": 0.139,
    "F": 0.084,
    "G": 0.054,
    "H": 0.006,
}
_VOA_NAME_TO_REGION = {
    "North East": "NORTH_EAST",
    "North West": "NORTH_WEST",
    "Yorkshire and The Humber": "YORKSHIRE",
    "East Midlands": "EAST_MIDLANDS",
    "West Midlands": "WEST_MIDLANDS",
    "East of England": "EAST_OF_ENGLAND",
    "London": "LONDON",
    "South East": "SOUTH_EAST",
    "South West": "SOUTH_WEST",
    "Wales": "WALES",
}


@lru_cache(maxsize=1)
def _download_workbook() -> openpyxl.Workbook:
    ref = load_config()["voa"]["council_tax"]
    r = requests.get(ref, headers=HEADERS, allow_redirects=True, timeout=60)
    r.raise_for_status()
    match = re.search(
        r'https://assets\.publishing\.service\.gov\.uk/media/[^"]+/[^"]+Summary_Tables\.xlsx',
        r.text,
    )
    if not match:
        raise ValueError("Could not find VOA council tax summary workbook link")
    workbook_url = match.group(0)
    workbook_response = requests.get(
        workbook_url,
        headers=HEADERS,
        allow_redirects=True,
        timeout=60,
    )
    workbook_response.raise_for_status()
    return openpyxl.load_workbook(BytesIO(workbook_response.content), data_only=True)


def _load_xlsx_response(response: requests.Response) -> openpyxl.Workbook:
    response.raise_for_status()
    try:
        return openpyxl.load_workbook(BytesIO(response.content), data_only=True)
    except BadZipFile as error:
        content_type = response.headers.get("content-type", "")
        preview = response.content[:160].decode("utf-8", errors="replace")
        preview = " ".join(preview.split())
        raise ValueError(
            "Expected an XLSX workbook from "
            f"{response.url}; got content-type {content_type!r} "
            f"with {len(response.content)} bytes: {preview!r}"
        ) from error


def _scotland_workbook_urls_from_page() -> list[str]:
    r = requests.get(
        _SCOTLAND_REF,
        headers=HEADERS,
        allow_redirects=True,
        timeout=60,
    )
    r.raise_for_status()
    urls: list[str] = []
    for match in _SCOTLAND_WORKBOOK_LINK.finditer(r.text):
        url = urljoin(_SCOTLAND_REF, unescape(match.group(1)))
        if url not in urls:
            urls.append(url)
    return urls


@lru_cache(maxsize=1)
def _download_scotland_workbook() -> openpyxl.Workbook:
    errors: list[str] = []

    def try_url(url: str) -> openpyxl.Workbook | None:
        for attempt in range(3):
            try:
                r = requests.get(
                    url,
                    headers=HEADERS,
                    allow_redirects=True,
                    timeout=60,
                )
                return _load_xlsx_response(r)
            except Exception as error:
                errors.append(f"{url}: {error}")
                if attempt < 2:
                    time.sleep(2**attempt)
        return None

    workbook = try_url(_SCOTLAND_WORKBOOK_URL)
    if workbook is not None:
        return workbook

    for url in _scotland_workbook_urls_from_page():
        workbook = try_url(url)
        if workbook is not None:
            return workbook

    raise ValueError(
        "Could not download Scotland council tax workbook. Recent errors: "
        + " | ".join(errors[-3:])
    )


def _to_float(value) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _fallback_scotland_band_counts() -> dict[str, float]:
    counts = {
        band: _SCOTLAND_FALLBACK_TOTAL_2025 * share
        for band, share in _SCOTLAND_FALLBACK_BAND_SHARES_2024_25.items()
    }
    counts["Total"] = _SCOTLAND_FALLBACK_TOTAL_2025
    return counts


def _get_scotland_band_counts() -> dict[str, float]:
    try:
        scotland_ws = _download_scotland_workbook()["Chargeable Dwellings 2025"]
    except Exception as error:
        logger.warning(
            "Using fallback Scotland council tax band distribution: %s",
            error,
        )
        return _fallback_scotland_band_counts()

    scotland_row = 8
    scotland_col_index = {
        "A": 2,
        "B": 3,
        "C": 4,
        "D": 5,
        "E": 6,
        "F": 7,
        "G": 8,
        "H": 9,
        "Total": 10,
    }
    return {
        band: _to_float(
            scotland_ws.cell(
                row=scotland_row,
                column=scotland_col_index[band],
            ).value
        )
        for band in [*_BANDS, "Total"]
    }


def get_targets() -> list[Target]:
    """Build council tax band targets from the latest VOA workbook."""
    wb = _download_workbook()
    ws = wb[_SHEET]
    ref = load_config()["voa"]["council_tax"]
    targets = []
    year = 2025

    headers = [ws.cell(row=_HEADER_ROW, column=col).value for col in range(1, 15)]
    header_index = {header: idx + 1 for idx, header in enumerate(headers)}

    for row_num in range(_HEADER_ROW + 1, ws.max_row + 1):
        geography = ws.cell(
            row=row_num, column=header_index["Geography [note 1]"]
        ).value
        if geography not in ("REGL", "NATL"):
            continue
        area_name = ws.cell(row=row_num, column=header_index["ONS area name"]).value
        region = _VOA_NAME_TO_REGION.get(area_name)
        if not region:
            continue
        for band in _BANDS:
            targets.append(
                Target(
                    name=f"voa/council_tax/{region}/{band}",
                    variable="council_tax_band",
                    source="voa",
                    unit=Unit.COUNT,
                    geographic_level=GeographicLevel.REGION,
                    geo_name=region,
                    values={
                        year: _to_float(
                            ws.cell(row=row_num, column=header_index[band]).value
                        )
                    },
                    is_count=True,
                    reference_url=ref,
                )
            )
        targets.append(
            Target(
                name=f"voa/council_tax/{region}/total",
                variable="council_tax_band",
                source="voa",
                unit=Unit.COUNT,
                geographic_level=GeographicLevel.REGION,
                geo_name=region,
                values={
                    year: _to_float(
                        ws.cell(
                            row=row_num,
                            column=header_index["All properties"],
                        ).value
                    )
                },
                is_count=True,
                reference_url=ref,
            )
        )

    scotland_band_counts = _get_scotland_band_counts()
    for band in _BANDS:
        targets.append(
            Target(
                name=f"voa/council_tax/SCOTLAND/{band}",
                variable="council_tax_band",
                source="voa",
                unit=Unit.COUNT,
                geographic_level=GeographicLevel.REGION,
                geo_name="SCOTLAND",
                values={year: scotland_band_counts[band]},
                is_count=True,
                reference_url=_SCOTLAND_REF,
            )
        )
    targets.append(
        Target(
            name="voa/council_tax/SCOTLAND/total",
            variable="council_tax_band",
            source="voa",
            unit=Unit.COUNT,
            geographic_level=GeographicLevel.REGION,
            geo_name="SCOTLAND",
            values={year: scotland_band_counts["Total"]},
            is_count=True,
            reference_url=_SCOTLAND_REF,
        )
    )

    return targets
