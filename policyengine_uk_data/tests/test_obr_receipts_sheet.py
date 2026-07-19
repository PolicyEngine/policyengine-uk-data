"""Tests for locating the OBR current-receipts table.

EFO releases renumber their worksheets between vintages. The cash-basis
receipts rows (NICs, VAT, fuel duties, capital gains tax, SDLT) were read from
a hard-coded sheet "3.9", but in the March 2026 tables current receipts sit on
3.8 and 3.9 is the APD forecast. Every lookup therefore raised, was caught, and
logged a warning — so all five targets vanished from the calibration surface
without failing anything.

These tests pin the title-based lookup and the loud failure that replaces the
silent drop.
"""

import openpyxl
import pytest

from policyengine_uk_data.targets.sources.obr import (
    _find_receipts_sheet,
    _parse_receipts,
)


def _wb(sheets: dict[str, str]) -> openpyxl.Workbook:
    """Build a workbook whose sheets carry their EFO title in cell B2."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, title in sheets.items():
        ws = wb.create_sheet(name)
        ws["B2"] = title
    return wb


def test_finds_receipts_sheet_by_title_not_number():
    """The March 2026 layout: receipts on 3.8, APD on 3.9."""
    wb = _wb(
        {
            "3.8": "3.8 Current receipts (on a cash basis)",
            "3.9": "3.9 APD forecast - projection of passenger numbers",
        }
    )
    assert _find_receipts_sheet(wb).title == "3.8"


def test_finds_receipts_sheet_after_renumbering():
    """A future vintage may move the table again; the title still finds it."""
    wb = _wb(
        {
            "3.9": "3.9 Some other forecast",
            "3.11": "3.11 Current receipts (on a cash basis)",
        }
    )
    assert _find_receipts_sheet(wb).title == "3.11"


def test_prefers_cash_basis_when_title_is_ambiguous():
    """Several sheets mention receipts; the cash-basis one is the right table."""
    wb = _wb(
        {
            "3.4": "3.4 Current receipts (accrued basis)",
            "3.8": "3.8 Current receipts (on a cash basis)",
        }
    )
    assert _find_receipts_sheet(wb).title == "3.8"


def test_raises_rather_than_guessing_when_ambiguous():
    """Two indistinguishable candidates must fail loudly, not pick the first."""
    wb = _wb(
        {
            "3.8": "3.8 Current receipts (one)",
            "3.9": "3.9 Current receipts (two)",
        }
    )
    with pytest.raises(ValueError, match="Refusing to guess"):
        _find_receipts_sheet(wb)


def test_raises_when_no_receipts_sheet():
    """Missing the table must fail loudly rather than yield zero targets."""
    wb = _wb({"3.9": "3.9 APD forecast"})
    with pytest.raises(ValueError, match="Current receipts"):
        _find_receipts_sheet(wb)


def _receipts_wb(*, populate: bool) -> openpyxl.Workbook:
    """A current-receipts sheet carrying every label, optionally with values."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # 3.4 (income tax, accrued basis) is read unconditionally; a missing sheet
    # raises KeyError rather than the ValueError the parser catches.
    wb.create_sheet("3.4")["B2"] = "3.4 Income tax (accrued basis)"
    ws = wb.create_sheet("3.8")
    ws["B2"] = "3.8 Current receipts (on a cash basis)"
    labels = [
        "National insurance contributions",
        "Value added tax",
        "Fuel duties",
        "Capital gains tax",
        "Stamp duty land tax",
    ]
    for offset, label in enumerate(labels):
        row = 10 + offset
        ws[f"B{row}"] = label
        if populate:
            for col in ("D", "E", "F", "G", "H", "I", "J"):
                ws[f"{col}{row}"] = 1.0
    return wb


def test_raises_when_rows_are_found_but_yield_no_values(monkeypatch):
    """A column shift finds every label and still produces nothing.

    The guard must key on targets produced, not on exceptions caught, or this
    reproduces the silent drop by a different route.
    """
    monkeypatch.setattr(
        "policyengine_uk_data.targets.sources.obr.load_config",
        lambda: {"obr": {"vintage": "test", "efo_receipts": "https://example.invalid"}},
    )
    with pytest.raises(ValueError, match="produced values"):
        _parse_receipts(_receipts_wb(populate=False))


def test_parses_rows_when_values_are_present(monkeypatch):
    """The same sheet with values yields one target per cash-basis row."""
    monkeypatch.setattr(
        "policyengine_uk_data.targets.sources.obr.load_config",
        lambda: {"obr": {"vintage": "test", "efo_receipts": "https://example.invalid"}},
    )
    targets = _parse_receipts(_receipts_wb(populate=True))
    names = {target.name for target in targets}
    assert "obr/capital_gains_tax" in names
    assert len(names) == 5
