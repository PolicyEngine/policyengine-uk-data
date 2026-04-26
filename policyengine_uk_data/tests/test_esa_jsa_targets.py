"""Tests for ESA/JSA calibration target definitions."""

import openpyxl

from policyengine_uk_data.targets.sources import dwp, obr


def test_dwp_esa_jsa_claimant_targets_exist():
    targets = {target.name: target for target in dwp.get_targets()}

    assert targets["dwp/esa_claimants"].variable == "esa"
    assert targets["dwp/esa_claimants"].values[2025] == 999_000

    assert targets["dwp/esa_contrib_claimants"].variable == "esa_contrib"
    assert targets["dwp/esa_contrib_claimants"].values[2025] == 620_000

    assert targets["dwp/esa_income_claimants"].variable == "esa_income"
    assert targets["dwp/esa_income_claimants"].values[2025] == 180_000

    # The official 2025 JSA caseload is New Style / contributory JSA.
    assert targets["dwp/jsa_claimants"].variable == "jsa_contrib"
    assert targets["dwp/jsa_claimants"].values[2025] == 71_000


def test_obr_jobseekers_allowance_maps_to_total_jsa():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "4.9"
    ws["B2"] = "Jobseeker's allowance"
    ws["C2"] = 0.2
    ws["D2"] = 0.3
    ws["E2"] = 0.2
    ws["F2"] = 0.2
    ws["G2"] = 0.3
    ws["H2"] = 0.2
    ws["I2"] = 0.2

    targets = {target.name: target for target in obr._parse_welfare(wb)}

    assert targets["obr/jobseekers_allowance"].variable == "jsa"
