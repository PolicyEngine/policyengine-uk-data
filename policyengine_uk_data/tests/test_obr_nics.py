"""Regression tests for OBR NIC target parsing.

Bug-hunt finding U8: ``_parse_nics`` originally only covered Class 1
employee and employer NICs, omitting Classes 2 (self-employed flat-rate),
3 (voluntary) and 4 (self-employed profit-based). Calibration had no
target for those receipts and pushed self-employment income downward to
compensate.

Follow-up (issue #378, partial close of #88):

- **Class 3** is voluntary contributions paid by people topping up
  their state-pension record. PolicyEngine UK exposes ``ni_class_3``
  as an input variable, but no dataset builder, imputation, or utility
  populates it — so a Class 3 target would be a flat-zero matrix
  column with no signal. The parser therefore drops it.
- **Classes 2 + 4** are bundled in recent OBR EFOs (e.g. March 2026)
  as a single "Class 4 and Class 2 Self employed NICs" line. The
  parser emits a combined ``obr/ni_self_employed`` target that uses
  ``custom_compute`` to sum the two PE-UK variables. If a future EFO
  reverts to separate rows, the legacy ``ni_class_2`` / ``ni_class_4``
  candidates still match and emit individually.
"""

from __future__ import annotations

import importlib.util
from unittest.mock import patch

import pytest

if importlib.util.find_spec("openpyxl") is None:
    pytest.skip("openpyxl not installed", allow_module_level=True)

import openpyxl  # noqa: E402

_FAKE_CONFIG = {
    "obr": {
        "vintage": "test",
        "efo_receipts": "https://example.invalid/receipts",
        "efo_expenditure": "https://example.invalid/expenditure",
    }
}


def _populate_sheet(ws, rows: list[tuple[str, list[float]]]) -> None:
    """Write rows of (label, fy_values) into column B / C-I of ``ws``."""
    for row_idx, (label, values) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=2, value=label)
        for col_idx, value in enumerate(values, start=3):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _build_combined_efo_workbook() -> openpyxl.Workbook:
    """Mimic the March 2026+ EFO Table 3.4 layout with combined SE NICs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3.4"
    _populate_sheet(
        ws,
        [
            ("Class 1 Employee NICs", [120.0] * 7),
            ("Class 1 Employer NICs", [140.0] * 7),
            ("Class 4 and Class 2 Self employed NICs", [4.8] * 7),
            # Class 3 is bundled into "Other NIC" in current EFOs and is
            # not directly extractable; included here only to confirm the
            # parser does not emit a Class 3 target.
            ("Other NIC", [0.5] * 7),
        ],
    )
    return wb


def _build_separate_efo_workbook() -> openpyxl.Workbook:
    """Mimic an older EFO layout with separate Class 2 / Class 4 rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3.4"
    _populate_sheet(
        ws,
        [
            ("Class 1 Employee NICs", [120.0] * 7),
            ("Class 1 Employer NICs", [140.0] * 7),
            ("Class 2 NICs", [0.3] * 7),
            ("Class 3 NICs", [0.05] * 7),
            ("Class 4 NICs", [4.5] * 7),
        ],
    )
    return wb


def test_parse_nics_combined_self_employed_line():
    """Current EFO format: combined Class 2+4 line gets a single target with
    custom_compute, alongside Class 1 employee/employer."""
    from policyengine_uk_data.targets.sources import obr as obr_module

    with patch.object(obr_module, "load_config", return_value=_FAKE_CONFIG):
        targets = obr_module._parse_nics(_build_combined_efo_workbook())

    names = {t.name for t in targets}
    assert names == {
        "obr/ni_employee",
        "obr/ni_employer",
        "obr/ni_self_employed",
    }

    combined = next(t for t in targets if t.name == "obr/ni_self_employed")
    assert combined.variable == "ni_self_employed"
    assert combined.values[2024] == pytest.approx(4.8e9)
    # The combined target is virtual — it must carry a custom_compute that
    # sums the two underlying PE-UK variables. Without it the loss matrix
    # falls through to the simple-GBP path and looks for a non-existent
    # `ni_self_employed` PE-UK variable.
    assert callable(combined.custom_compute)


def test_parse_nics_falls_back_to_separate_classes_for_old_efo():
    """If a future EFO publishes Class 2 and Class 4 on separate rows again,
    the legacy candidates pick them up individually."""
    from policyengine_uk_data.targets.sources import obr as obr_module

    with patch.object(obr_module, "load_config", return_value=_FAKE_CONFIG):
        targets = obr_module._parse_nics(_build_separate_efo_workbook())

    names = {t.name for t in targets}
    assert names == {
        "obr/ni_employee",
        "obr/ni_employer",
        "obr/ni_class_2",
        "obr/ni_class_4",
    }

    class_4_target = next(t for t in targets if t.variable == "ni_class_4")
    assert class_4_target.values[2024] == pytest.approx(4.5e9)


def test_parse_nics_intentionally_skips_class_3_in_combined_efo():
    """In the current combined-line layout, Class 3 is hidden inside
    "Other NIC" and the parser must not emit it (#378)."""
    from policyengine_uk_data.targets.sources import obr as obr_module

    with patch.object(obr_module, "load_config", return_value=_FAKE_CONFIG):
        targets = obr_module._parse_nics(_build_combined_efo_workbook())

    assert "obr/ni_class_3" not in {t.name for t in targets}
    assert "ni_class_3" not in {t.variable for t in targets}


def test_parse_nics_intentionally_skips_class_3_in_separate_efo():
    """Even when an EFO publishes a Class 3 row, the parser must not emit
    a target for it — the underlying PE-UK variable has no signal (#378)."""
    from policyengine_uk_data.targets.sources import obr as obr_module

    with patch.object(obr_module, "load_config", return_value=_FAKE_CONFIG):
        targets = obr_module._parse_nics(_build_separate_efo_workbook())

    assert "obr/ni_class_3" not in {t.name for t in targets}
    assert "ni_class_3" not in {t.variable for t in targets}


def test_parse_nics_tolerates_alt_label_wording():
    """Common wording variants (`Self-Employed` vs `self-employed`) match."""
    from policyengine_uk_data.targets.sources import obr as obr_module

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3.4"
    _populate_sheet(
        ws,
        [
            ("Class 1 Employee NICs", [100.0] * 7),
            ("Class 1 Employer NICs", [110.0] * 7),
            ("Class 2 Self-Employed NICs", [0.2] * 7),
            ("Class 3 Voluntary NICs", [0.04] * 7),
            ("Class 4 Self-Employed NICs", [4.0] * 7),
        ],
    )

    with patch.object(obr_module, "load_config", return_value=_FAKE_CONFIG):
        targets = obr_module._parse_nics(wb)

    assert {t.variable for t in targets} == {
        "ni_employee",
        "ni_employer",
        "ni_class_2",
        "ni_class_4",
    }
