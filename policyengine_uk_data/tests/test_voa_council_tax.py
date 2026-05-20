from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest

from policyengine_uk_data.targets.sources import voa_council_tax


class DummyResponse:
    def __init__(
        self,
        content: bytes = b"",
        *,
        text: str = "",
        url: str = "https://example.test/file.xlsx",
        content_type: str = "application/octet-stream",
    ):
        self.content = content
        self.text = text
        self.url = url
        self.headers = {"content-type": content_type}

    def raise_for_status(self):
        return None


def _workbook_bytes() -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Chargeable Dwellings 2025"
    worksheet.cell(row=8, column=10).value = 2_623_149
    out = BytesIO()
    workbook.save(out)
    return out.getvalue()


def test_scotland_workbook_download_falls_back_to_source_page(monkeypatch):
    voa_council_tax._download_scotland_workbook.cache_clear()
    calls = []

    def fake_get(url, **kwargs):
        calls.append(url)
        if len(calls) <= 3:
            return DummyResponse(
                b"<html>temporary bad response</html>",
                url=url,
                content_type="text/html",
            )
        if url == voa_council_tax._SCOTLAND_REF:
            return DummyResponse(
                text=(
                    '<a href="/binaries/content/documents/govscot/publications/'
                    "statistics/2019/04/council-tax-datasets/documents/"
                    "number-of-chargeable-dwellings/"
                    "chargeable-dwellings---september-2025-data/"
                    "chargeable-dwellings---september-2025-data/"
                    "govscot%3Adocument/"
                    "CTAXBASE%2B2025%2B-%2BTables%2B-%2B"
                    'Chargeable%2BDwellings.xlsx">download</a>'
                ),
                url=url,
                content_type="text/html",
            )
        return DummyResponse(
            _workbook_bytes(),
            url=url,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    monkeypatch.setattr(voa_council_tax.requests, "get", fake_get)
    monkeypatch.setattr(voa_council_tax.time, "sleep", lambda _: None)

    workbook = voa_council_tax._download_scotland_workbook()

    assert workbook["Chargeable Dwellings 2025"].cell(row=8, column=10).value == (
        2_623_149
    )
    assert calls[:3] == [voa_council_tax._SCOTLAND_WORKBOOK_URL] * 3
    assert voa_council_tax._SCOTLAND_REF in calls


def test_load_xlsx_response_reports_bad_content_type():
    with pytest.raises(ValueError, match="Expected an XLSX workbook"):
        voa_council_tax._load_xlsx_response(
            DummyResponse(
                b"<html>not a workbook</html>",
                content_type="text/html",
            )
        )
